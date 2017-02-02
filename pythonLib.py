from bs4 import BeautifulSoup
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import re
import itertools
from selenium.webdriver.common.keys import Keys


## Input & Output file names, please change it accordingly
OUTPUT_FILENAME = 'Out_put_combined_template_new.xlsx'
INPUT_FILENAME = 'Data_Collection_Config.xlsx'

## File directory, please change it accordingly
#root = 'C:/Users/mglobe10/Desktop/YIfan test/'
#root_out = 'F:/Data_collection_system/Main Folder/Street Account US & Euro/'
#root_in = 'F:/Data_collection_system/'
#root_pandas = 'F:/Data_collection_system\Main Folder\Street Account US & Euro\Out_put_combined_template.xlsx'
#root_out = '/Users/pengjiawei/PycharmProjects/project1/'
#root_in = '/Users/pengjiawei/PycharmProjects/project1/'
#root_pandas = '/Users/pengjiawei/PycharmProjects/project1/Out_put_combined_template_new.xlsx'
root_out = '/Users/huangxingyue/Desktop/'
root_in = '/Users/huangxingyue/Desktop/'
root_pandas = '/Users/huangxingyue/Desktop/Out_put_combined_template_new.xlsx'


## keywords to signal a buy action:
buy_wordbag = ['upgrade', 'initiated with a buy', 'initiated with an accumulate', \
            'initiated with an overweight', 'initiated with a positive', 'initiated outperform', \
            'initiated buy', 'initiated overweight', 're-initiated accumulate', \
            'reinstated buy', 're-initiated buy', 'reiterated buy', 'assumed buy', \
            'initiated add', 'resumed outperform', 'initiated buy', 'initiated accumulate', \
            'estimates raised', 'initiated positive', 'added to focus list', 'added to conviction buy list',\
            'initiated sector outperform', 'initiated sector outperformer', 'assumed outperform', 'initiated market outperform', 'resumed buy']

## keywords to signal a sell action:
sell_wordbag = ['downgrade', 'initiated with a sell', 'initiated with a reduce', \
            'initiated with an underweight', 'initiated with a negative', 'initiated underperform', \
            'initiated sell', 'initiated underweight', 'reinstated sell', 'reinstated underweight', \
            'reinstated underperform', 'initiated reduce', 'removed from conviction buy list',\
            'removed from focus list', 'assumed sell', 'cautious comments', 'removed from conviction buy list']  

## keywords to signal a newtral rating:
kickout_wordbag = ['initiated hold', 'initiated market perform', 'initiated neutral', \
            'initiated with hold', 'initiated with a hold', 'initiated sector perform', \
            'assumed neutral', 'resumed neutral', 'reinstated neutral', 'reinstated equal-weight', \
            'reinstated hold', 'initiated equal-weight', 'resumed equal-weight',\
            'initiated sector weight', 'initiated secotr perform', 'target decreased', \
            'target increased', 'resumed hold', 'price target cut', 'price target lowered', 'price target raised', 'reiterated overweight', 'initiation with neutral']  

##  Combine keywords together for buy_or_sell as input.
wordbag = buy_wordbag + sell_wordbag + kickout_wordbag

 
##  Helper function to click element in a webpage
def click(driver,elem):
        actions = ActionChains(driver)
        actions.move_to_element(elem)
        actions.click(elem)
        actions.perform()



## Return country name
def get_country_name(symbol):
        find_dot = re.search("\.",symbol)
        if find_dot:
                symbol_list = symbol.split(".")
                
                if symbol_list[-1]=="NA": 
                        country = "Netherlands"
                elif symbol_list[-1]=="AU":
                        country = "Australia"
                elif symbol_list[-1]=="AV":
                        country = "Austria"
                elif symbol_list[-1]=="BB":
                        country = "Belgium"
                elif symbol_list[-1]=="BZ":
                        country = "Brazil"
                elif symbol_list[-1]=="CN":
                        country = "Canada"
                elif symbol_list[-1]=="CI":
                        country = "Chile"
                elif symbol_list[-1]=="CH":
                        country = "China"
                elif symbol_list[-1]=="CB":
                        country = "Colombia"
                elif symbol_list[-1]=="CY":
                        country = "Cyprus"
                elif symbol_list[-1]=="CP":
                        country = "Czech Republic"
                elif symbol_list[-1]=="DC":
                        country = "Denmark"
                elif symbol_list[-1]=="FO":
                        country = "Faroe Islands"
                elif symbol_list[-1]=="FH":
                        country = "Finland"
                elif symbol_list[-1]=="FP":
                        country = "France"
                elif symbol_list[-1]=="GR":
                        country = "Germany"
                elif symbol_list[-1]=="GA":
                        country = "Greece"
                elif symbol_list[-1]=="HK":
                        country = "Hong Kong"
                elif symbol_list[-1]=="HB":
                        country = "Hungary"
                elif symbol_list[-1]=="ID":
                        country = "Ireland"
                elif symbol_list[-1]=="IM":
                        country = "Italy"
                elif symbol_list[-1]=="IN":
                        country = "India"
                elif symbol_list[-1]=="IT":
                        country = "Israel"
                elif symbol_list[-1]=="KS":
                        country = "South Korea"
                elif symbol_list[-1]=="KZ":
                        country = "Kazakhstan"
                elif symbol_list[-1]=="JP":
                        country = "Japan"
                elif symbol_list[-1]=="LN":
                        country = "London"
                elif symbol_list[-1]=="LX":
                        country = "Luxembourg"
                elif symbol_list[-1]=="MM":
                        country = "Mexico"
                elif symbol_list[-1]=="MK":
                        country = "Malaysia"
                elif symbol_list[-1]=="NO":
                        country = "Norway"
                elif symbol_list[-1]=="NZ":
                        country = "New Zealand"
                elif symbol_list[-1]=="NG":
                        country = "Nigeria"
                elif symbol_list[-1]=="PM":
                        country = "Philippines"
                elif symbol_list[-1]=="PL":
                        country = "Portugal"
                elif symbol_list[-1]=="PW":
                        country = "Poland"
                elif symbol_list[-1]=="RU":
                        country = "Russia"
                elif symbol_list[-1]=="SP":
                        country = "Singapore"
                elif symbol_list[-1]=="SL":
                        country = "Slovenia"
                elif symbol_list[-1]=="SM":
                        country = "Spain"
                elif symbol_list[-1]=="SS":
                        country = "Sweden"
                elif symbol_list[-1]=="VX":
                        country = "Switzerland"
                elif symbol_list[-1]=="SW":
                        country = "Switzerland"
                elif symbol_list[-1]=="SJ":
                        country = "South Africa"
                elif symbol_list[-1]=="TT":
                        country = "Taiwan"
                elif symbol_list[-1]=="TB":
                        country = "Thailand"
                elif symbol_list[-1]=="TI":
                        country = "Turkey"                
                else: 
                        country = "Country list needs to expand"
        else:
                country = "United States"
        return country



## Helper function to write each record into output file
def insert_record_sheet_data(input_list):
        '''
        input parameter is a list consist of 25 elements
            0) get_date                                      col 1 for Amr and col 1 for global
            1) get_time                                      col 2 for Amr and col 2 for global
            2) get_country_code                              col 3 for Amr and col 6 for global
            3) country_name                                  col 4 for Amr and col 8 for global
            4) get_detail                                    col 5 for Amr and col 12 for global
            5) buy_or_sell                                   col 6 for Amr and col 13 for global
        '''
        print('inside inserting data in Record sheet')
        main_file = root_out + OUTPUT_FILENAME

        ## Open the excel bood
        wb = openpyxl.load_workbook(main_file)

        ## Locate the right sheet
        if input_list[3] == 'United States':
            done_sh = wb.get_sheet_by_name('Amr Ratings')
            print ("Sheet Amr Ratings")
        else:
            done_sh = wb.get_sheet_by_name('Global Ratings')
            print ("Sheet Global Ratings")

        ## Locate the blank row appended at the bottom
        max_row = done_sh.get_highest_row()
        #max_col = done_sh.get_highest_column()

        print(max_row)
        #next is added by Margaret in June .21th
        #set the date-format in excel to get rid of the green corner in cells, this makes the excel_reader parse dates format successfully

        ## Write in the data!
        if input_list[3] == 'United States':
            done_sh.cell(row = (max_row+1), column = 1).value = input_list[0]
            done_sh.cell(row = (max_row+1), column = 2).value = input_list[1]
            done_sh.cell(row = (max_row+1), column = 3).value = input_list[2]
            done_sh.cell(row = (max_row+1), column = 4).value = input_list[3]
            done_sh.cell(row = (max_row+1), column = 5).value = input_list[4]
            done_sh.cell(row = (max_row+1), column = 6).value = input_list[5]

        else:
            done_sh.cell(row = (max_row+1), column = 1).value = input_list[0]
            done_sh.cell(row = (max_row+1), column = 2).value = input_list[1]
            done_sh.cell(row = (max_row+1), column = 6).value = input_list[2]
            done_sh.cell(row = (max_row+1), column = 8).value = input_list[3]
            done_sh.cell(row = (max_row+1), column = 12).value = input_list[4]
            done_sh.cell(row = (max_row+1), column = 13).value = input_list[5]

        wb.save(main_file)




## Helper function to specify buy/sell
def buy_or_sell(detail):
        '''
        input: a string (news headline) named detail
        
        output: 4 categories of 'buy'/'sell'/'kickout'/' ' 
        '''
        result = ''
        #def bag_parse(str_1): return str_1 in detail.lower() ## whether a pattern in a string

        ## anonymous function to extract all matchings of keyword in the input string
        def bag_parse(str_1): return re.findall(str_1, detail) ## return a list of matched key words

        ## Using anonymous function bag_parse to match all keywors in wordbag whether 
        ## appeared in the input string        
        Sub_list = list(itertools.chain.from_iterable(map(bag_parse, wordbag))) ## collapse list of list of key words

        ## If matched keywors is a subset of any categorized wordbag, return the category name
        if set(Sub_list).issubset(buy_wordbag) and len(Sub_list) != 0:
                result = 'buy'
        elif set(Sub_list).issubset(sell_wordbag) and len(Sub_list) != 0:
                result = 'sell'
        elif set(Sub_list).issubset(kickout_wordbag) and len(Sub_list) != 0:
                result = 'kickout'
        return result



## Helper function for extracting the time, ticker and headline for each record
## Called inside the main while loop after selecting the Up/Downgrade tab and calendar date
def ind_page(driver):
        
        print("inside ind Page")

        driver.implicitly_wait(10)
        get_date = driver.find_element_by_xpath("//*[@id='headlineDate']").text.strip()
        print("get_date",get_date)
        if '(Archives)' in get_date:
                get_date=get_date.replace('(Archives)','').strip()
        #cur_date = get_date
##        get_main_div = driver.find_element_by_xpath("//*[@id='content']")
        soup = BeautifulSoup(driver.page_source,"html5lib")
        main_div = soup.find("div",{"id":"content"})
        
##        try:
##                get_lbl_div=soup.find("div",{"class":"MHB_Labels"})

        ## Collect all records to be extracted, and use For loop to iterate each of them        
        get_all_divs_inside=main_div.find_all("div",{"id":re.compile("story\d{7}")})

        ## Print out how many records in total of current webpage (Do Not Include Sub-bullet)
        print("Length is : ",len(get_all_divs_inside))

        ## Enter into the for loop to extract record one by one
        for each_record in get_all_divs_inside:
##                        each_record.location_once_scrolled_into_view
                ## Print out record ID first, sub-bullets share a commmon record ID
                print(each_record['id'],"id is ")
                ## Extract news release time here
				
                try:
                        get_time=each_record.find("td",{"class":"sa-date-time"}).get_text().strip()
                except:
                        try:
                                get_time=each_record.find("td",{"class":"tabletimered"}).get_text().strip()
                        except:
                                get_time = ''
                                print ('current time: ', get_time)
             
                ## Extract ticker for this record, If no ticker detected, leave it blank
                ## For sub bullet news, there are no tickers here!! We use other methods later
                try:
                        get_country_code=each_record.find("a",{"class":"sa-ticker"}).get_text().strip()
                        country_name = get_country_name(get_country_code)
                except:
                        country_name = ' '
                        get_country_code = ' '
                        print ('current country code: ', get_country_code)
                        
                ## create an internal list, each element is a list of ticker, company name and headline combo (list of list)
                ## In case of sub bullet news, each element for one news
                Inter_list = []
                
                ## If ticker successfully extracted, meaning no sub bullet in this records, 
                ## we extract the head line in regular way:
                if get_country_code != ' ':
                        try:
                                get_detail=each_record.find("span",{"class":"sa-headline"}).get_text().strip()
                        except:
                                try:
                                        get_detail=each_record.find("span",{"class":"tabletextRed"}).get_text().strip()
                                except:
                                        get_detail = ' '                    
                        ## And put this record in the list created above
                        Inter_list.append([get_country_code, country_name, get_detail])
                
                ## If no ticker extracted so far, meaning there are multiple sub bullet points in this record
                ## Then detecting ticker within each sub bullet according to certain parterns
                ## Also, there may be multiple tickers within on sub-bullet point
                else:
                        ## Extract all sub bullet points into a list, and use a for loop
                        ## to iterate each of them
                        sub_test = each_record.find("span",{"class":"sa-comment-text"}).get_text().split('\n')
                        for i in range(len(sub_test)): # extract patterns may be appeared in headline
                                # Pattern1: _ABCD_ eg ..._GOOG123_...
                                code_list1 = re.findall("\s[A-Z0-9]+\s", sub_test[i])
                                for k in range(len(code_list1)):
                                        code_list1[k] = code_list1[k][1:-1] ## delete the white spaces
                                # Pattern2: AB2C3.C.CB  eg: ...GOOG123.B.CN...
                                code_list2 = re.findall("[A-Z0-9]+\.[A-Z]\.[A-Z]+", sub_test[i])
                                # Pattern3: (AB3C4.CB   eg: (GOOG123.CN...
                                code_list3 = re.findall("\([A-Z0-9]+\.[A-Z][A-Z]", sub_test[i])
                                for k in range(len(code_list3)):
                                        code_list3[k] = code_list3[k][1:] ## delete the parenthese
                                # Pattern4: (ACS2_  eg: (GOOG123_...
                                code_list4 = re.findall("\([A-Z0-9]+\s", sub_test[i])
                                for k in range(len(code_list4)):
                                        code_list4[k] = code_list4[k][1:-1]  ## delete the parenthese
                                # Pattern5: (ABC2) eg: (GOOG123)
                                code_list5 = re.findall("\([A-Z0-9]+\)", sub_test[i])
                                for k in range(len(code_list5)):
                                        code_list5[k] = code_list5[k][1:-1]  ## delete the parenthese

                                ## if we already have detected pattern5,4,3 or 2, then ignore pattern1
                                if len(code_list4) != 0 or len(code_list3) != 0 or len(code_list2) != 0 \
                                    or len(code_list5) != 0:
                                        code_list1 = []

                                ## Combine all tetected ticker together
                                code_list = code_list1 + code_list2 + code_list3 + code_list4 \
                                            + code_list5 ## extract all possible tickers
                                print (code_list)

                                ## If only only one ticker detected, extract corresponding headline
                                if len(code_list) == 1:
                                        #get_country_code == code_list[0]
                                        country_name = get_country_name(code_list[0])
                                        get_detail = sub_test[i]
                                        ## And added it into the internal list
                                        Inter_list.append([code_list[0], country_name, get_detail])
                                ## If no tickers detected in sub bullets, then leave it blank
                                elif  len(code_list) == 0:                                     
                                        get_country_code = ' '                                                
                                        country_name = ' '  
                                        get_detail = ' '
                                        ## And added it into the internal list 
                                        Inter_list.append([get_country_code, country_name, get_detail])
                                ## If multiple tickers detected, 
                                ##  extract corresponding headlines one by one using for loop below:
                                else:
                                        for j in range(len(code_list)):
                                            #get_country_code == code_list[j]
                                            country_name = get_country_name(code_list[j])
                                            get_detail = sub_test[i]
                                            ## And added it into the internal list one by one
                                            Inter_list.append([code_list[j], country_name, get_detail])                                            

                                  

                ## loop to iterate internal list to incert each record into excel sheet
                for i in range(len(Inter_list)):
                        print(Inter_list[i][1],"country_name")
                        print(get_time,Inter_list[i][0],Inter_list[i][2],"get_time,get_country_code,get_detail")
                        
                        ## Only records with a valid ticker code will be inserted into excel sheet
                        ## and Call buy_or_sell function in pythonLib.py to kick out neutral rated headlines
                        if Inter_list[i][0] != ' ' and  buy_or_sell(Inter_list[i][2]) != 'kickout':

                                ## Call buy_or_sell function in pythonLib.py to categorize buy/sell
                                get_buysell = buy_or_sell(Inter_list[i][2])

                                ## Combined all to be inserted information into a list to comply with
                                ## the format needed for next function          
                                #get_date = datetime.strptime(get_time, "%d-%b-%y")
                                input_list = [get_date,get_time,Inter_list[i][0],Inter_list[i][1],Inter_list[i][2]] \
                                            + [get_buysell]                      

                                ## Call insert_record_sheet_data function in pythonLib.py to write in excel sheet
                                ## Use error handler to protect file from corrupted
                                try:
                                        insert_record_sheet_data(input_list)    
                                except Exception as e:
                                        print  (str(e) + ' ' + 'Failed in writing ' + str(get_date) + ' ' + str(get_time) + ' ' + str(Inter_list[i][0]))

        ## Records for this calender date has been inserted.                                                    
        return driver




## Helper function to convert Month format
def check_month(month_before):
        month=''
        if(month_before=='jan' or month_before=='january'):
            month='01'
        elif(month_before=='feb' or month_before=='february'):
            month='02'
        elif(month_before=='mar' or month_before=='march'):
            month='03'
        elif(month_before=='apr' or month_before=='april'):
            month='04'
        elif(month_before=='may' or month_before=='may'):
            month='05'
        elif(month_before=='jun' or month_before=='june'):
            month='06'
        elif(month_before=='jul' or month_before=='july'):
            month='07'
        elif(month_before=='aug' or month_before=='august'):
            month='08'
        elif(month_before=='sep' or month_before=='september'):
            month='09'
        elif(month_before=='oct' or month_before=='october'):
            month='10'
        elif(month_before=='nov' or month_before=='november'):
            month='11'
        elif(month_before=='dec' or month_before=='december'):
            month='12'
        return month




## Helper function to extract start date and end date from input file
def get_dates():
    main_file = root_in + INPUT_FILENAME
    print (main_file)
    wb = openpyxl.load_workbook(main_file)
    #wb = openpyxl.load_workbook(main_file, use_iterators=True)
    login_sh = wb.get_sheet_by_name('Street Account Data')
    '''
        login_rows = login_sh.iter_rows()
        next(login_rows)
        for row_get in login_rows:
                start_date_xl  = row_get[2].value
                end_date_xl=row_get[3].value
                break
    '''
    start_date_xl = login_sh['C2'].value
    end_date_xl = login_sh['D2'].value
    login_sh['D40'].value = ''
    wb.save(main_file)
    return (start_date_xl,end_date_xl)




## Helper function to extract credentials from input file        
def get_user_info():
    main_file = root_in + INPUT_FILENAME
    #wb = openpyxl.load_workbook(main_file, use_iterators=True)
    wb = openpyxl.load_workbook(main_file)
    login_sh = wb.get_sheet_by_name('Street Account Data')
    '''
        login_rows = login_sh.iter_rows()
        next(login_rows)
        for row_get in login_rows:
                user_name_xl  = row_get[0].value
                password_xl=row_get[1].value
                break
    '''
    user_name_xl = login_sh['A2'].value
    password_xl = login_sh['B2'].value
    login_sh['D40'].value = ''
    wb.save(main_file)
    return user_name_xl,password_xl




## Helper function to log in the webpage   
def login(driver):
        user_name_xl,password_xl=get_user_info()
        print(user_name_xl,password_xl,"user_name_xl,password_xl")

        user_name = driver.find_element_by_xpath("//*[@id='ctl00_tbxUsername']")
        user_name.click()
        user_name.clear()
        user_name.send_keys(user_name_xl)
        pw = driver.find_element_by_xpath("//*[@id='ctl00_tbxPassword']")
        pw.click()
        pw.clear()
        pw.send_keys(password_xl)
        pw.send_keys(Keys.RETURN)
        driver.implicitly_wait(10)
        return driver



## Dictionary for mapping stock exchange code
Exc_map = {'Argentina': ['BCBA'],
           'Mexico': ['BMV'],
           'Brazil': ['BVMF'],
           'Canada': ['CNSX', 'CVE', 'TSE'],
           'United States': ['NASDAQ', 'NYSE', 'NYSEARCA', 'NYSEMKT', 'OPRA', 'OTCBB', 'OTCMKTS'],
           'Netherlands': ['AMS'],
           'Italy': ['BIT'],
           'Spain': ['BME'],
           'Denmark': ['CPH'],
           'Belgium': ['EBR'],
           'Portugal': ['ELI'],
           'France': ['EPA'],
           'Germany': ['ETR', 'FRA'],
           'Finland': ['HEL'],
           'Iceland': ['ICE'],
           'Turkey': ['IST'],
           'London': ['LON'],
           'Russia': ['MCX'],
           'Latvia': ['RSE'],
           'Sweden': ['STO'],
           'Switzerland': ['SWX', 'VTX'],
           'Estonia': ['TAL'],
           'Austria': ['VIE'],
           'Lithuania': ['VSE'],
           'Poland': ['WSE'],
           'South Africa': ['JSE'],
           'Saudi Arabia': ['TADAWUL'],
           'Israel': ['TLV'],
           'Thailand': ['BKK'],
           'India': ['BOM', 'NSE'],
           'Malaysia': ['KLSE'],
           'Hong Kong': ['HKG'],
           'Indonesia': ['IDX'],
           'South Korea': ['KOSDAQ', 'KRX'],
           'Singapore': ['SGX'],
           'China': ['SHA', 'SHE'],
           'Taiwan': ['TPE'],
           'Japan': ['TYO'],
           'Australia': ['ASX'],
           'New Zealand': ['NZE']}


## Dictionary for mapping stock exchange code
Exc_map_2 = {'Argentina': ['', 'BCBA:'],
             'Mexico': ['', 'BMV:'],
             'Brazil': ['', 'BVMF:'],
             'Canada': ['', 'CNSX:', 'CVE:', 'TSE:'],
             'United States': ['', 'NASDAQ:', 'NYSE:', 'NYSEARCA:', 'NYSEMKT:', 'OPRA:', 'OTCBB:', 'OTCMKTS:'],
             'Netherlands': ['', 'AMS:'],
             'Italy': ['', 'BIT:'],
             'Spain': ['', 'BME:'],
             'Denmark': ['', 'CPH:'],
             'Belgium': ['', 'EBR:'],
             'Portugal': ['', 'ELI:'],
             'France': ['', 'EPA:'],
             'Germany': ['', 'ETR:', 'FRA:'],
             'Finland': ['', 'HEL:'],
             'Iceland': ['', 'ICE:'],
             'Turkey': ['', 'IST:'],
             'London': ['', 'LON:'],
             'Russia': ['', 'MCX:'],
             'Latvia': ['', 'RSE:'],
             'Sweden': ['', 'STO:'],
             'Switzerland': ['', 'SWX:', 'VTX:'],
             'Estonia': ['', 'TAL:'],
             'Austria': ['', 'VIE:'],
             'Lithuania': ['', 'VSE:'],
             'Poland': ['', 'WSE:'],
             'South Africa': ['', 'JSE:'],
             'Saudi Arabia': ['', 'TADAWUL:'],
             'Israel': ['', 'TLV:'],
             'Thailand': ['', 'BKK:'],
             'India': ['', 'BOM:', 'NSE:'],
             'Malaysia': ['', 'KLSE:'],
             'Hong Kong': ['', 'HKG:'],
             'Indonesia': ['', 'IDX:'],
             'South Korea': ['', 'KOSDAQ:', 'KRX:'],
             'Singapore': ['', 'SGX:'],
             'China': ['', 'SHA:', 'SHE:'],
             'Taiwan': ['', 'TPE:'],
             'Japan': ['', 'TYO:'],
             'Australia': ['', 'ASX:'],
             'New Zealand': ['', 'NZE:']}



## Dictionary for mapping stock exchange code from yahoo finance
Exc_map_3 = {'Argentina': ['', '.BA'],
             'Mexico': ['', '.MX'],
             'Brazil': ['', '.SA'],
             'Canada': ['', '.TO', '.V'],
             'United States': ['', '.CBT', '.CME', '.NYB', '.CMX', '.NYM', '.OB', '.PK'],
             'Netherlands': ['', '.AS'],
             'Italy': ['', '.MI'],
             'Spain': ['', '.BC', '.BI', '.MF', '.MC', '.MA'],
             'Denmark': ['', '.CO'],
             'Belgium': [''],
             'Portugal': [''],
             'France': ['', '.NX', '.PA'],
             'Germany': ['', '.DE', '.BE', '.BM', '.DU', '.F', '.HM', '.HA', '.MU', '.SG'],
             'Finland': [''],
             'Iceland': [''],
             'Turkey': [''],
             'London': ['', '.L'],
             'Russia': [''],
             'Latvia': [''],
             'Sweden': ['', '.ST'],
             'Switzerland': ['', '.SW'],
             'Estonia': [''],
             'Austria': ['', '.VI'],
             'Lithuania': [''],
             'Poland': [''],
             'South Africa': [''],
             'Saudi Arabia': [''],
             'Israel': ['', '.TA'],
             'Thailand': [''],
             'India': ['', '.BO', '.NS'],
             'Malaysia': [''],
             'Hong Kong': ['', '.HK'],
             'Indonesia': ['', '.JK'],
             'South Korea': ['', '.KS', '.KQ'],
             'Singapore': ['', '.SI'],
             'China': ['', '.SS', '.SZ'],
             'Taiwan': ['', '.TWO', '.TW'],
             'Japan': [''],
             'Australia': ['', '.AX'],
             'New Zealand': ['', '.NZ'],
             'Chile': ['', '.SN'],
             'Norway': ['', '.OL']}