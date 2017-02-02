import pandas as pd
import numpy as np
import urllib.request as urllib2
import datetime as dt
import dateutil.relativedelta
import openpyxl
import sys
from bs4 import BeautifulSoup
import pythonLib
import requests
import time


# Helper function to get historical daily volume data (raw data)
# Exchange Parameter added, applies to Non-US stock
# Calculate average 10 days volume, median 10 days volume and median 3 month volume
def get_volume_data(symbol, date):
    """
    input: ticker, end date

    output: average 10 days volume, median 10 days volume and median 3 month volume
    """

    # find corresponding exchange code
    exc_list = pythonLib.Exc_map_3.get(pythonLib.get_country_name(symbol), [''])

    bench_vol = 0.0
    result = []
    ticker = symbol.split('.')[0]
    print('Getting volume data for: ' + ticker)
    # start date is 3 month before end date
    start_date = dt.datetime.strptime(date, "%d-%b-%y") - dateutil.relativedelta.relativedelta(months=4)
    start_date = start_date.strftime('%m %d %Y')  # convert date format...
    start_date = start_date.split(' ')
    start_date = '&a=' + start_date[0] + '&b=' + start_date[1] + '&c=' + start_date[2]
    end_date = dt.datetime.strptime(date, "%d-%b-%y") - dateutil.relativedelta.relativedelta(months=1)
    end_date = end_date.strftime('%m %d %Y')
    end_date = end_date.split(' ')
    end_date = '&d=' + end_date[0] + '&e=' + end_date[1] + '&f=' + end_date[2]

    for i in range(len(exc_list)):
        url_root = 'http://chart.finance.yahoo.com/table.csv?s='
        url_root += ticker + exc_list[i] + start_date + end_date + '&g=d&ignore=.csv'
        # check whether the url is valid
        try:
            df = pd.read_csv(url_root)
            print(url_root)
        except:
            continue
        # get the volume data
        vol = df['Volume']
        # check whether has enough data
        if len(vol) <= 10:
            print("Not enough data")
            continue
        avg_10d_vol = np.mean(vol[:10])
        med_10d_vol = np.median(vol[:10])
        if len(vol) <= 60:
            print("Not enough data")
            med_3m_vol = 'null'
        else:
            med_3m_vol = np.median(vol)
        if avg_10d_vol > bench_vol:
            bench_vol = avg_10d_vol
            result.append((avg_10d_vol, med_10d_vol, med_3m_vol))

    if not result:
        return 'null', 'null', 'null'
    else:
        return result[-1]


# Helper function to get historical daily volume data (raw data)
# Do NOT consider exchange, applies to US stock
# Calculate average 10 days volume, median 10 days volume and median 3 month volume
def get_volume_data_2(symbol, date):
    """
    input: ticker, end date

    output: average 10 days volume, median 10 days volume and median 3 month volume
    """

    ticker = symbol.split('.')[0]
    print('Getting volume data for: ' + ticker)
    # start date is 3 month before end date
    start_date = dt.datetime.strptime(date, "%d-%b-%y") - dateutil.relativedelta.relativedelta(months=4)
    start_date = start_date.strftime('%m %d %Y')  # convert date format...
    start_date = start_date.split(' ')
    start_date = '&a=' + start_date[0] + '&b=' + start_date[1] + '&c=' + start_date[2]
    end_date = dt.datetime.strptime(date, "%d-%b-%y") - dateutil.relativedelta.relativedelta(months=1)
    end_date = end_date.strftime('%m %d %Y')
    end_date = end_date.split(' ')
    end_date = '&d=' + end_date[0] + '&e=' + end_date[1] + '&f=' + end_date[2]

    url_root = 'http://chart.finance.yahoo.com/table.csv?s='
    url_root += ticker + start_date + end_date + '&g=d&ignore=.csv'
    print(url_root)
    # check whether the url is valid
    try:
        df = pd.read_csv(url_root)
    except:
        return 'null', 'null', 'null'
    # get the volume data
    vol = df['Volume']
    # check whether has enough data
    if len(vol) <= 10:
        print("Not enough data")
        return 'null', 'null', 'null'
    avg_10d_vol = np.mean(vol[:10])
    med_10d_vol = np.median(vol[:10])
    if len(vol) <= 60:
        print("Not enough data")
        return avg_10d_vol, med_10d_vol, 'null'
    med_3m_vol = np.median(vol)

    return avg_10d_vol, med_10d_vol, med_3m_vol


# Function to get Beta and average 3 month volume
# Exchange Parameter added, applies to Non-US stock
def beta_avg_3m_vol(symbol):
    """
    input: ticker

    output: beta and average 3month volume
    """

    exc_list = pythonLib.Exc_map_3.get(pythonLib.get_country_name(symbol), [''])
    result = []
    bench_vol = 0.0
    ticker = symbol.split('.')[0]
    print('Getting beta for: ' + ticker)

    for i in range(len(exc_list)):
        url_root = 'http://finance.yahoo.com/quote/'
        url_root += ticker + exc_list[i] + '/?p=' + ticker + exc_list[i]
        status = False
        trytime = 0
        page = None
        # check whether the url is valid
        while not status and trytime <= 5:
            try:
                # page = urllib2.urlopen(url_root, timeout=10)
                page = requests.get(url_root, timeout=30)
                print(url_root)
                status = True
            except (requests.HTTPError, requests.ConnectionError):
                break
            except:
                trytime += 1
                print('timeout')
                time.sleep(60)
        if not page:
            continue
        # c = page.read().decode('utf-8')
        c = page.text
        soup = BeautifulSoup(c, "html5lib")
        page.close()
        table = soup.findAll('table')
        headings = [td.get_text() for i in table for td in i.find_all('td')]
        if 'Avg Vol (3m)' in headings:
            vol_index = headings.index('Avg Vol (3m)')
            avg_vol_3m = headings[vol_index + 1]
            if avg_vol_3m == 'N/A':
                avg_vol_3m = 'null'
            else:
                avg_vol_3m = ''.join(avg_vol_3m.split(','))
        else:
            avg_vol_3m = 'null'
        if 'Beta' in headings:
            beta_index = headings.index('Beta')
            beta = headings[beta_index + 1]
            if beta == 'N/A':
                beta = 'null'
        else:
            beta = 'null'
        if avg_vol_3m == 'null' and bench_vol == 0.0:
                result.append((avg_vol_3m, beta))
        elif avg_vol_3m != 'null' and float(avg_vol_3m) > bench_vol:
            bench_vol = float(avg_vol_3m)
            result.append((avg_vol_3m, beta))
    return result[-1]


# Function to get Beta and average 3 month volume
# Do NOT consider exchange, applies to US stock
def beta_avg_3m_vol_2(symbol):
    """
    input: ticker

    output: beta and average 3month volume
    """

    ticker = symbol.split('.')[0]
    print('Getting beta for: ' + ticker)
    url_root = 'http://finance.yahoo.com/quote/'
    url_root += ticker + '/?p=' + ticker
    print(url_root)

    status = False
    trytime = 0
    page = None
    # check whether the url is valid
    while not status and trytime <= 5:
        try:
            # page = urllib2.urlopen(url_root, timeout=10)
            page = requests.get(url_root, timeout=30)
            status = True
        except (requests.HTTPError, requests.ConnectionError):
            return 'null' 'null'
        except:
            trytime += 1
            print('timeout')
            time.sleep(60)
    if not page:
        return 'null' 'null'
    # c = page.read().decode('utf-8')
    c = page.text
    soup = BeautifulSoup(c, "html5lib")
    page.close()
    table = soup.findAll('table')
    headings = [td.get_text() for i in table for td in i.find_all('td')]
    if 'Avg Vol (3m)' in headings:
        vol_index = headings.index('Avg Vol (3m)')
        avg_vol_3m = headings[vol_index + 1]
        if avg_vol_3m == 'N/A':
            avg_vol_3m = 'null'
    else:
        avg_vol_3m = 'null'
    if 'Beta' in headings:
        beta_index = headings.index('Beta')
        beta = headings[beta_index + 1]
        if beta == 'N/A':
            beta = 'null'
    else:
        beta = 'null'

    return avg_vol_3m, beta


# Function to get Sector Classification of a Company
# Exchange Parameter added, applies to Non-US stock
def get_sector(symbol):
    """
    input: ticker

    output: Company's sector
    """

    exc_list = pythonLib.Exc_map_2.get(pythonLib.get_country_name(symbol), [''])
    ticker = symbol.split('.')[0]
    print('Getting sector for: ' + ticker)
    sector = []

    for i in range(len(exc_list)):
        url_root = 'https://www.google.com/finance?q=' + exc_list[i] + '%3A' + ticker
        # check whether the url is valid
        status = False
        trytime = 0
        page = None
        # check whether the url is valid
        while not status and trytime <= 5:
            try:
                # page = urllib2.urlopen(url_root, timeout=10)
                page = requests.get(url_root, timeout=30)
                print(url_root)
                status = True
            except (requests.HTTPError, requests.ConnectionError):
                break
            except:
                trytime += 1
                print('timeout')
                time.sleep(60)
        if not page:
            continue
        # c = page.read().decode('utf-8')
        c = page.text
        soup = BeautifulSoup(c, "html5lib")
        page.close()
        a = soup.findAll('a')
        sec = [j.get_text() for j in a if j.get('id') == 'sector']
        if sec:
            sector.append(sec[0])

    sector.append('null')
    return sector[0]


# Function to get Sector Classification of a Company
# Do NOT consider exchange, applies to US stock
def get_sector_2(symbol):
    """
    input: ticker

    output: Company's sector
    """

    ticker = symbol.split('.')[0]
    print('Getting sector for: ' + ticker)

    url_root = 'https://www.google.com/finance?q=' + ticker
    print(url_root)

    status = False
    trytime = 0
    page = None
    # check whether the url is valid
    while not status and trytime <= 5:
        try:
            # page = urllib2.urlopen(url_root, timeout=10)
            page = requests.get(url_root, timeout=30)
            status = True
        except (requests.HTTPError, requests.ConnectionError):
            return 'null' 'null'
        except:
            trytime += 1
            print('timeout')
            time.sleep(60)
    if not page:
        return 'null'
    # c = page.read().decode('utf-8')
    c = page.text
    soup = BeautifulSoup(c, "html5lib")
    page.close()
    a = soup.findAll('a')
    sector = [j.get_text() for j in a if j.get('id') == 'sector']
    sector.append('null')
    return sector[0]


# Function to get pre/post earnings announcement company's ticker and time
def get_ear_df(date):
    """
    input: date

    output: dataframe containing ticker and time
    """
    l = []
    date = dt.datetime.strptime(date, "%d-%b-%y").strftime("%Y%m%d")
    url_root = 'https://biz.yahoo.com/research/earncal/' + date + '.html'
    print('Getting pre/post earnings')
    print(url_root)
    status = False
    trytime = 0
    page = None
    # check whether the url is valid
    while not status and trytime <= 5:
        try:
            # page = urllib2.urlopen(url_root, timeout=10)
            page = requests.get(url_root, timeout=30)
            status = True
        except (requests.HTTPError, requests.ConnectionError):
            return pd.DataFrame()
        except:
            trytime += 1
            print('timeout')
            time.sleep(60)
    if not page:
        return pd.DataFrame()
    # c = page.read().decode('utf-8')
    c = page.text
    soup = BeautifulSoup(c, "html5lib")
    page.close()
    table = soup.findAll('table')[6]
    tr = table.findAll('tr')
    for i in range(2, len(tr) - 1):
        td = tr[i].findAll('td')
        l.append([td[1].get_text(), td[2].get_text()])
    df = pd.DataFrame(data=l, columns=['Symbol', 'Time'])

    return df


# Function to check whether have pre/post earnings
# Exchange Parameter added, applies to Non-US stock
def get_ear(ticker, df):
    """
    input: ticker
           data frame containing ticker and time
    output: pre/post earning
    """
    exc_list = pythonLib.Exc_map_3.get(pythonLib.get_country_name(ticker), [''])
    for i in range(len(exc_list)):
        symbol = ticker + exc_list[i]
        if symbol in df['Symbol'].values:
            time = df[df['Symbol'] == symbol]['Time'].values[0]
            return time

    return 'null'


# Function to check whether have pre/post earnings
# Do NOT consider exchange, applies to US stock
def get_ear_2(ticker, df):
    """
    input: ticker
           data frame containing ticker and time
    output: pre/post earning
    """
    if ticker in df['Symbol'].values:
        time = df[df['Symbol'] == ticker]['Time'].values[0]
        return time
    return 'null'


# Function to convert list_of_dates into list_of_rows
def conv_dates_to_rows(path_file, sheet_name, list_of_dates):
    """
    input: output file name (directory convention a little bit different)
           name of target sheet
           list of dates to be converted
    output: of list of row numbers (Type: Long)
    """
    df = pd.read_excel(path_file, sheet_name)

    bool_list = []

    for my_date in list_of_dates:
        if not bool_list:
            bool_list = (df['US Date'] == my_date).tolist()
        else:
            bool_list = np.logical_or(bool_list, (df['US Date'] == my_date).tolist()).tolist()

    my_list = df[bool_list].index.tolist()
    my_list = (np.array(my_list) + 2).tolist()

    return my_list


# Function to write data into the file
def insert_data_sheet_date(list_of_dates):
    """
    input: date of current row
    """
    print('Inserting data into file:')
    main_file = pythonLib.root_out + pythonLib.OUTPUT_FILENAME
    wb = openpyxl.load_workbook(main_file)
    done_sh_Amr = wb.get_sheet_by_name('Amr Ratings')
    done_sh_Glo = wb.get_sheet_by_name('Global Ratings')

    list_of_rows_Amr = conv_dates_to_rows(pythonLib.root_pandas, 'Amr Ratings', list_of_dates)
    list_of_rows_Glo = conv_dates_to_rows(pythonLib.root_pandas, 'Global Ratings', list_of_dates)
    list_to_use = []

    # get data
    for done_sh in [done_sh_Amr, done_sh_Glo]:

        if done_sh == done_sh_Amr:
            Ticker_col = 'C'
            list_to_use = list_of_rows_Amr
        elif done_sh == done_sh_Glo:
            Ticker_col = 'F'
            list_to_use = list_of_rows_Glo

        for index in list_to_use:

            stock_code = str(done_sh[Ticker_col + str(index)].value)
            get_date = str(done_sh['A' + str(index)].value)

            print('(Stock: ' + stock_code + ', Date: ' + get_date + ')')

            df = get_ear_df(get_date)

            # get data and write data into excel
            if Ticker_col == 'C':
                try:
                    avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data_2(stock_code, get_date)
                    avg_3m_vol, beta = beta_avg_3m_vol_2(stock_code)
                except ValueError:
                    print(stock_code + ':ValueError')
                    try:
                        avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data_2(stock_code, get_date)
                        avg_3m_vol, beta = beta_avg_3m_vol_2(stock_code)
                    except ValueError:
                        print(stock_code + ':ValueError')
                        try:
                            avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data_2(stock_code, get_date)
                            avg_3m_vol, beta = beta_avg_3m_vol_2(stock_code)
                        except ValueError:
                            print(stock_code + ':ValueError')
                            continue
                sector = get_sector_2(stock_code)
                earning = get_ear_2(stock_code, df)
                done_sh.cell(row=index, column=26).value = sector
                done_sh.cell(row=index, column=27).value = avg_3m_vol
                done_sh.cell(row=index, column=28).value = avg_10d_vol
                done_sh.cell(row=index, column=29).value = med_10d_vol
                done_sh.cell(row=index, column=30).value = med_3m_vol
                done_sh.cell(row=index, column=31).value = earning
                done_sh.cell(row=index, column=32).value = beta

            elif Ticker_col == 'F':
                try:
                    avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data(stock_code, get_date)
                    avg_3m_vol, beta = beta_avg_3m_vol(stock_code)
                except ValueError:
                    print(stock_code + ':ValueError')
                    try:
                        avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data(stock_code, get_date)
                        avg_3m_vol, beta = beta_avg_3m_vol(stock_code)
                    except ValueError:
                        print(stock_code + ':ValueError')
                        try:
                            avg_10d_vol, med_10d_vol, med_3m_vol = get_volume_data(stock_code, get_date)
                            avg_3m_vol, beta = beta_avg_3m_vol(stock_code)
                        except ValueError:
                            print(stock_code + ':ValueError')
                            continue
                sector = get_sector(stock_code)
                earning = get_ear(stock_code, df)
                done_sh.cell(row=index, column=33).value = sector
                done_sh.cell(row=index, column=34).value = avg_3m_vol
                done_sh.cell(row=index, column=35).value = avg_10d_vol
                done_sh.cell(row=index, column=36).value = med_10d_vol
                done_sh.cell(row=index, column=37).value = med_3m_vol
                done_sh.cell(row=index, column=38).value = earning
                done_sh.cell(row=index, column=39).value = beta

            list_to_use = []

    wb.save(main_file)


# Read in the start date and end date in input file by calling get_dates function
start_date_xl, end_date_xl = pythonLib.get_dates()
start_date_xl = start_date_xl.date()
end_date_xl = end_date_xl.date()

# Main function of this script
# Running the while loop when the dates are valid
if start_date_xl <= end_date_xl:

    start = str(start_date_xl.strftime('%d-%b-%y'))
    end = str(end_date_xl.strftime('%d-%b-%y'))

    Date_list = []

    for My_date in pd.date_range(start, end):
        Date_list.append(str(My_date.date().strftime('%d-%b-%y')))

    try:
        # Call helper function to write into excel
        insert_data_sheet_date(Date_list)
        print("Completed")
    except Exception as e:
        print("Cannot write into excel",)
        print(e, sys.exc_info())
