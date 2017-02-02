import openpyxl
from urllib.request import urlopen, URLError
import datetime
import time
import re
import pandas as pd
import os
import csv
import xlwt
import pickle
from openpyxl.workbook import Workbook


# set the function to get the data from google finance
def rest():
    readwb = openpyxl.load_workbook('/Users/pengjiawei/Desktop/m_global/config/Price & Volume config_NEW.xlsx')
    sheetimport = readwb.get_sheet_by_name(name='US Intraday Data')
    resttime = float(sheetimport.cell('B10').value)
    time.sleep(resttime)


def pause():
    print(time.localtime())
    rest()
    print(time.localtime())


def IntradayPrice(ticker, date_str, time_str, datalist):

    if len(datalist) < 9:
        print('Bad Ticker!')  # no data received
        return 'NULL'

    # change the date and time index format in order to compare with the ones on the website
    date_str_list = date_str.split('/')
    time_str_list = time_str.split(':')
    idate = datetime.date(int(date_str_list[2]), int(date_str_list[0]), int(date_str_list[1]))
    itime = datetime.time(int(time_str_list[0]), int(time_str_list[1]))
    dt = datetime.datetime.combine(idate, itime)
    dtstamp = int(time.mktime(dt.timetuple()))
    if source == 'G':
        try:
            tmpstr = datalist[7].split(',')[0]  # 7th line is the data line,starting with 'a' to indicate a new date;
            EarlistTimeStamp = int(tmpstr[1:len(tmpstr)])  # the first date stamp in the data received
            data = datalist[7:len(datalist)]
        except:
            tmpstr = datalist[8].split(',')[0]  # 8th line is the data line,starting with 'a' to indicate a new date;
            EarlistTimeStamp = int(tmpstr[1:len(tmpstr)])  # the first date stamp in the data received
            data = datalist[8:len(datalist)]
    elif source == 'Y':
        for i in range(len(datalist)):
            if 'volume:' in datalist[i]:
                break
        EarlistTimeStamp = int(datalist[i+1].split(',')[0])  # the first date stamp in the data received
        data = datalist[i+1:len(datalist)]
        EarlistTimeStamp = str(datetime.datetime.fromtimestamp(EarlistTimeStamp))
        EarlistTimeStamp = EarlistTimeStamp[:11] + '9:30:00'
        EarlistTimeStamp = int(time.mktime(datetime.datetime.strptime(EarlistTimeStamp, '%Y-%m-%d %H:%M:%S').timetuple()))


    # get data according to the index
    if int(dtstamp) < EarlistTimeStamp:
        print('DateTime exceeds the date range.')
        return 'NULL'

    else:
        print('data exists!')
        for str1 in data:
            count = 0
            str_list = str1.split(',')
            if source == 'G':
                if 'a' in str_list[0]:
                    DateStamp = int(str_list[0][1:len(str_list[0])])
                    MinStamp = DateStamp
                else:
                    try:
                        MinStamp = DateStamp + 60 * int(str_list[0])
                    except ValueError:
                        MinStamp = DateStamp
                if MinStamp == dtstamp:
                    print(str_list[1])
                    return float(str_list[1])
                elif dtstamp < MinStamp:
                    print(ticker+': cannot find data at current minute. Return open price of next minute.')
                    return float(str_list[4])
            elif source == 'Y':
                DateStamp = str(datetime.datetime.fromtimestamp(int(str_list[0])))
                DateStamp = DateStamp[:-2] + '00'
                DateStamp = int(time.mktime(datetime.datetime.strptime(DateStamp, '%Y-%m-%d %H:%M:%S').timetuple()))
                MinStamp = DateStamp
                if MinStamp == dtstamp:
                    print(str_list[1])
                    return float(str_list[1])
                elif dtstamp < MinStamp:
                    if str(idate) == '-'.join(str(datetime.datetime.today())[:10].split('-')):
                        print(ticker+': cannot find data at current minute. Return open price of next minute.')
                        return float(str_list[4])
                    else:
                        return 'NULL'
        print('DateTime exceeds the date range.')
        return 'NULL'


# pop out each column associated with pricing given specific sheet #########
# 'sheet' is index sheet we used to find the data from google finance
# 'sheet2' is where we save the new data
def SheetUpdate(sheet, sheet2, tradedate):

    # count the row numbers
    counter = 0
    m = 2
    while sheet.cell('A' + str(m)).value:
        counter += 1
        m += 1
    try:
        f = open(path_pickle + 'US.pickle', 'rb')
        start_num = pickle.load(f)
    except IOError:
        print("The process is starting from the first ticker. No need to continue from checkpoint")
        start_num = 2
    print("start_num is ", start_num)
    # get ticker
    for j in range(start_num, counter+2):  # counter+2
        Ticker = sheet.cell('A' + str(j)).value
        print(Ticker)
        Temp_ticker = Ticker.split(' ')
        if len(Temp_ticker) == 2:
            tickerG = Temp_ticker[0] + '.' + Temp_ticker[1]
        else:
            tickerG = Ticker

        # access to the link where we will download the data from
        # tickerG could be 'None', breaks re.search with "TypeError: expected string or buffer"
        if tickerG is None:
            tickerG = ''  # change it to empty string
        elif re.search('-', tickerG):
            tickerG = tickerG.replace('-', '.')  # ticker is changed if it contains dash
        # determine the source of data
        if source == 'G':
            url = 'https://www.google.com/finance/getprices?i=60&p=10d&f=d,o,h,l,c,v&df=cpct&q=' + tickerG
        elif source == 'Y':
            if tradedate == ''.join(str(datetime.datetime.today())[:10].split('-')):
                print('current date, get 1min data')
                url = 'http://chartapi.finance.yahoo.com/instrument/1.0/' + tickerG + '/chartdata;type=quote;range=1d/csv'
            else:
                url = 'http://chartapi.finance.yahoo.com/instrument/1.0/' + tickerG + '/chartdata;type=quote;range=10d/csv'
                print('not current date, get 5min data')

        try:
            html = urlopen(url).read().decode()
            if source == 'Y' and j % float(frequency) == 0:
                pause()
        except URLError as e:
            if hasattr(e, 'reason'):
                print('We failed to reach a server.')
                print('Reason: ', e.reason)
            elif hasattr(e, 'code'):
                print('The server could not fulfill the request.')
                print('Error code: ', e.code)
            pause()
            try:
                html = urlopen(url).read().decode()
            except URLError as e:
                if hasattr(e, 'reason'):
                    print('We failed to reach a server.')
                    print('Reason: ', e.reason)
                elif hasattr(e, 'code'):
                    print('The server could not fulfill the request.')
                    print('Error code: ', e.code)
                pause()
                html = 'EXCHANGE%3DUNKNOWN+EXCHANGE\nMARKET_OPEN_MINUTE=570\nMARKET_CLOSE_MINUTE=960\nINTERVAL=60\n' \
                       'COLUMNS=DATE,CLOSE,HIGH,LOW,OPEN,VOLUME\nDATA=\n'

        datalist = html.splitlines()

        # get tradedate and change it to the useful format
        tradedatei = datetime.datetime.strptime(tradedate, '%Y%m%d').strftime('%m/%d/%Y')  # format change for curdate
        print(tradedate)
        # call the function and pop out
        # save when we finish each row
        for i in range(0, len(inputtimes)):
            openprice1 = IntradayPrice(Ticker, tradedatei, inputtimes[i], datalist)
            sheet2[j-1][4+i] = openprice1

        print(sheet2[j-1])

        with open(newpathfilecsv, "r", newline='') as csvfile:
            lines = csvfile.readlines()
            if len(lines) == counter + 1:
                continue
        csvfile.close()
        with open(newpathfilecsv, "a", newline='') as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(sheet2[j-1])
        pickle_counter = j + 1
        f = open(path_pickle + 'US.pickle', 'wb')
        pickle.dump(pickle_counter, f)
        f.close()



##################################################################################################
############################## I/O Staff #########################################################
##################################################################################################

# read the config file
path_root = os.getcwd()
path_pickle = '/Users/pengjiawei/Desktop/m_global/project3/pickle files/'
readwb = openpyxl.load_workbook('/Users/pengjiawei/Desktop/m_global/config/Price & Volume config_NEW.xlsx')
sheetimport = readwb.get_sheet_by_name(name='US Intraday Data')

# Read Input and check options from excel sheet
filename = sheetimport.cell('B2').value
filepath = sheetimport.cell('B3').value
pathfile = filepath + filename
newfilename = sheetimport.cell('B5').value
newfilepath = sheetimport.cell('B6').value
newfilenamecsv = sheetimport.cell('B13').value
newpathfile = newfilepath + newfilename
newpathfilecsv = newfilepath + newfilenamecsv
print(newpathfile)
transformcsv = sheetimport.cell('B14').value
source = sheetimport.cell('B15').value
frequency = sheetimport.cell('B11').value
wb1 = openpyxl.load_workbook(pathfile)

# create new excel file
wb2 = Workbook()

#####################  pop out the new file and sheets ###############################################

# create the sheets that need to be update in the new file
date1 = sheetimport.cell('B7').value
date2 = sheetimport.cell('B8').value
delta = date2 - date1
inputtime = sheetimport.cell('B9').value
intratime = sheetimport.cell('B12').value
inputtimes = []
startrow = 17
while sheetimport.cell('A'+str(startrow)).value:
    inputtimes.append(str(sheetimport.cell('A' + str(startrow)).value.hour) + ":" +
                      str(sheetimport.cell('A' + str(startrow)).value.minute))
    startrow += 1

if intratime != 1:
    # for i in range(0,delta.days+1):
    tradedate = date1 + datetime.timedelta(days=0)
    newname = tradedate.strftime("%Y%m%d")
    sheeticompare = wb1.get_sheet_by_name(name='Input')
    sheeti = []
    # write into the first row
    firstrow = ['Date', 'Time', 'Symbol', 'GVkey']

    for j in range(0, len(inputtimes)):
        firstrow.append(inputtimes[j])
    sheeti.append(firstrow)
    # write into the first four columns
    counteri = 1
    m = 2
    while sheeticompare.cell('A' + str(m)).value:
        counteri += 1
        m += 1
    for n in range(2, counteri + 1):
        row_to_add = [None] * (len(firstrow)+1)
        row_to_add[0:4] = [newname, inputtime, sheeticompare.cell('A'+str(n)).value]
        sheeti.append(row_to_add)
    # start to add data into the new sheet
    with open(newpathfilecsv, "a+", newline='') as csvfile:
        csvfile.seek(0)
        if not csvfile.read():
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(sheeti[0])
    csvfile.close()
    tradedate = tradedate.strftime("%Y%m%d")
    SheetUpdate(sheeticompare, sheeti, tradedate)

else:
    intratime = 2

'''
def SheetUpdate(sheet, sheet2, tradedate):

    # count the row numbers
    counter = 0
    m = 2
    while sheet.cell('B' + str(m)).value:
        counter += 1
        m += 1
    # get ticker
    for j in range(2, counter+2):
        Ticker = sheet.cell('B' + str(j)).value
        print(Ticker)

        Temp_ticker = Ticker.split(' ')
        if len(Temp_ticker) == 2:
            tickerG = Temp_ticker[0] + '.' + Temp_ticker[1]
        else:
            tickerG = Ticker

        # access to the link where we will download the data from
        if re.search('-', tickerG):
            tickerG = tickerG.replace('-', '.')  # ticker is changed if it contains dash
                # determine the source of data
        if source == 'G':
            url = 'https://www.google.com/finance/getprices?i=60&p=10d&f=d,o,h,l,c,v&df=cpct&q=' + tickerG
        elif source == 'Y':
            if tradedate == ''.join(str(datetime.datetime.today())[:10].split('-')):
                print('current date, get 1min data')
                url = 'http://chartapi.finance.yahoo.com/instrument/1.0/' + tickerG + '/chartdata;type=quote;range=1d/csv'
            else:
                url = 'http://chartapi.finance.yahoo.com/instrument/1.0/' + tickerG + '/chartdata;type=quote;range=10d/csv'
                print('not current date, get 5min data')

        try:
            html = urlopen(url).read().decode()
            if source == 'Y':
                pause()
        except URLError as e:
            if hasattr(e, 'reason'):
                print('We failed to reach a server.')
                print('Reason: ', e.reason)
            elif hasattr(e, 'code'):
                print('The server could not fulfill the request.')
                print('Error code: ', e.code)
            pause()
            try:
                html = urlopen(url).read().decode()
            except URLError as e:
                if hasattr(e, 'reason'):
                    print('We failed to reach a server.')
                    print('Reason: ', e.reason)
                elif hasattr(e, 'code'):
                    print('The server could not fulfill the request.')
                    print('Error code: ', e.code)
                pause()

                html = 'EXCHANGE%3DUNKNOWN+EXCHANGE\nMARKET_OPEN_MINUTE=570\nMARKET_CLOSE_MINUTE=960\nINTERVAL=60\n' \
                       'COLUMNS=DATE,CLOSE,HIGH,LOW,OPEN,VOLUME\nDATA=\n'

        datalist = html.splitlines()

        # get tradedate and change it to the useful format

        tradedatei = datetime.datetime.strptime(tradedate, '%Y%m%d').strftime('%m/%d/%Y')  # format change for curdate
        print(tradedate)
        # get time and change it to the useful format; the data will be popped out to new sheet from that time point
        time = sheetimport.cell('B9').value
        intratime = 2
        i0 = 571
        if time.minute // intratime == 0:
            initial = ((time.hour - 9) * 60 + (time.minute - 30))//intratime
        else:
            initial = (((time.hour - 9) * 60 + (time.minute - 30))//intratime) + 1
        if initial >= 0 and (initial < 389//intratime + 2):
            initial = initial
        else:
            initial = 0
        i0 += (intratime * initial)
    
        for i in range(initial, 389//intratime + 1):
            if i0 % 60 == 0 or i0 % 60 <= 10:
                Time = str(i0//60) + ':' + '0' + str(i0 % 60)
            else:
                Time = str(i0//60) + ':' + str(i0 % 60)

            # call the function and pop out
            # save when we finish each row
            openprice1 = IntradayPrice(Ticker, tradedatei, Time, datalist)
            sheet2.write(j-1, 4+i, openprice1)
            i0 += intratime
        for p in range(0, initial):
            sheet2.write(j-1, 4+p, 'N/A')
        j += 1
        wb2.save(newpathfile)

wb3 = xlwt.Workbook(style_compression=2)
filename = filename.replace('.', '(1).')
pathfile = filepath + filename
newfilename = newfilename.replace('.', '(1).')
newfilenamecsv = newfilenamecsv.replace('.', '(1).')
newpathfile = newfilepath + newfilename
newpathfilecsv = newfilepath + newfilenamecsv
for i in range(0, delta.days+1):
    tradedate = date1+datetime.timedelta(days=i)
    newname = tradedate.strftime("%Y%m%d")
    sheeticompare = wb1.get_sheet_by_name(name='Input')
    sheeti = wb3.add_sheet(newname)
    # write into the first row
    sheeti.write(0, 0, 'Date')
    sheeti.write(0, 1, 'Time')
    sheeti.write(0, 2, 'Symbol')
    sheeti.write(0, 3, 'GVkey')
    ii0 = 571
    for j in range(0, 389//intratime + 1):
        if ii0 % 60 == 0 or ii0 % 60 <= 10:
            Time = str(ii0 // 60) + ':' + '0' + str(ii0 % 60)
        else:
            Time = str(ii0 // 60) + ':' + str(ii0 % 60)
        sheeti.write(0, 4+j, Time)
        ii0 = ii0 + intratime

    #  write into the first four columns
    counteri = 1
    m = 2
    while sheeticompare.cell('A' + str(m)).value:
        counteri += 1
        m += 1
    for n in range(2,  counteri+1):
        style = xlwt.XFStyle()
        style.num_format_str = 'mm/dd/yy'
        sheeti.write(n-1, 0, newname, style)
        style = xlwt.XFStyle()
        style.num_format_str = 'h:mm:ss'
        sheeti.write(n-1, 1, inputtime, style)
        sheeti.write(n-1, 2, sheeticompare.cell('A'+str(n)).value)

    # start to add data into the new sheet
    tradedate = tradedate.strftime("%Y%m%d")
    SheetUpdate(sheeticompare, sheeti, tradedate)
    wb3.save(newpathfile)

wb3.save(newpathfile)
if transformcsv == 'Y':
    xls = pd.ExcelFile(newpathfile)
    df = xls.parse(newname, index_col=None, na_values=['NA'])
    df.to_csv(newpathfilecsv, sep='|')
'''


