import openpyxl
import pandas as pd
import csv
import sys
import datetime
import requests
from bs4 import BeautifulSoup
import re
import time
from urllib.request import urlopen


def url_open(url):
    status = False
    trytime = 0
    info1 = None
    while not status and trytime <= 5:
        try:
            print(url)
            info1 = urlopen(url)
            encoding = info1.info().get_content_charset('utf-8')
            info1 = info1.read().decode(encoding)
            status = True
        except:
            trytime += 1
            print('Can not access to the url. Try again after 1 minute.')
            print('It is the ' + str(trytime) + " trial.")
            time.sleep(60)
    return info1


def GetPriceURL(ticker):
    tickerG = ticker
    if re.search('-', tickerG):

        tickerG = tickerG.replace('-', '.')  # ticker is changed if it contains dash
    if source == 'G':
        url = 'https://www.google.com/finance/getprices?i=60&p=10d&f=d,o,h,l,c,v&df=cpct&q='+tickerG
    elif source == 'Y':
        url = 'http://chartapi.finance.yahoo.com/instrument/1.0/' + tickerG + '/chartdata;type=quote;range=1d/csv'
    html = url_open(url)
    try:
        datalist = html.splitlines()
    except:
        print("---- Can't access the page, probably blocked ----")
        datalist = "NULL"
    return datalist


def IntradayPrice(ticker, date_str, time_str, datalist):

    if len(datalist) < 8:
        print('Bad Ticker!')  # no data received
        return 'NULL'
    if source == 'G':
        if int(datalist[1][-3:]) != 570 or int(datalist[2][-3:]) != 960:
            print('Wrong Exchange Symbol!')
            return 'NULL'

    date_str_list = date_str.split('_')
    time_str_list = time_str.split(':')
    idate = datetime.date(int(date_str_list[0]), int(date_str_list[1]), int(date_str_list[2]))
    itime = datetime.time(int(time_str_list[0]), int(time_str_list[1]))
    dt = datetime.datetime.combine(idate, itime)
    dtstamp = time.mktime(dt.timetuple())

    if source == 'G':
        data = datalist[7:]
        tmpstr = datalist[7].split(',')[0]   # 7th line is the data line,starting with 'a' to indicate a new date;
        EarlistTimeStamp = float(tmpstr[1:len(tmpstr)])  # the first date stamp in the data received
        DateStamp = 0
        print(dtstamp, EarlistTimeStamp)
    elif source == 'Y':
        data = datalist[17:]
        EarlistTimeStamp = int(datalist[17].split(',')[0])  # the first date stamp in the data received
        EarlistTimeStamp = str(datetime.datetime.fromtimestamp(EarlistTimeStamp))
        EarlistTimeStamp = EarlistTimeStamp[:11] + '9:30:00'
        EarlistTimeStamp = int(time.mktime(datetime.datetime.strptime(EarlistTimeStamp, '%Y-%m-%d %H:%M:%S').timetuple()))
        print(dtstamp, EarlistTimeStamp)

    if dtstamp < EarlistTimeStamp:
        print('DateTime exceeds the date range.')
        return 'NULL'
    else:
        for str1 in data:
            str_list = str1.split(',')
            if source == 'G':
                if str1[0] == 'a':
                    DateStamp = float(str_list[0][1:len(str_list[0])])
                    MinStamp = DateStamp
                else:
                    try:
                        MinStamp = DateStamp + 60 * int(str_list[0])
                    except ValueError:
                        MinStamp = DateStamp

                if float(MinStamp) == dtstamp:
                    return float(str_list[1])
                elif dtstamp < float(MinStamp):
                    print(ticker+': cannot find data at current minute. Return open price of next minute.')
                    return float(str_list[4])
            elif source == 'Y':
                DateStamp = str(datetime.datetime.fromtimestamp(int(str_list[0])))
                DateStamp = DateStamp[:-2] + '00'
                DateStamp = int(time.mktime(datetime.datetime.strptime(DateStamp, '%Y-%m-%d %H:%M:%S').timetuple()))
                MinStamp = DateStamp
                if MinStamp == dtstamp:
                    print(str_list[1])
                    return float(str_list[1])
                elif dtstamp < MinStamp:
                    print(ticker+': cannot find data at current minute. Return open price of next minute.')
                    return float(str_list[4])
    return 'NULL'

print("reading config.xlsx")
path_config = "/Users/pengjiawei/Desktop/Data_Collection_Config.xlsx"
book_config = openpyxl.load_workbook(path_config)
sheet_config = book_config.get_sheet_by_name(name='Intraday_position_file')
startdate = sheet_config.cell("B1").value
enddate = sheet_config.cell("B2").value
path_output = sheet_config.cell("B11").value
name_output = sheet_config.cell("B10").value
division = float(sheet_config.cell("B7").value)
NAV = float(sheet_config.cell("B6").value)
cap = float(sheet_config.cell("B8").value)
cap_ipo = sheet_config.cell("B9").value
source = sheet_config.cell("B12").value
if startdate > enddate:
    sys.exit("Enddate should be later then the Startdate.. ")
if datetime.datetime.now() > startdate + datetime.timedelta(days=10):
    sys.exit("Startdate should be witthin 10 days form now!")

print("reading headline.xlsx")
path_headline = sheet_config.cell("B4").value
book_headline = openpyxl.load_workbook(path_headline)
sheet_headline = book_headline.get_sheet_by_name(name="Ratings Change")

print("reading equities.csv")
path_equities = sheet_config.cell("B5").value
with open(path_equities) as csvfile:
    equities = csv.reader(csvfile)
    equities = [x[0].split("|") for x in equities]

open_time = datetime.time(9, 30, 0)
end_time = datetime.time(16, 0, 0)
dates = pd.date_range(startdate, enddate)
for date in dates:
    print("now process date: ", date)
    data = [[None, "Symbol", "Signal", "Side", "Weight", "Time", "AvgVol3m"]]
    count = 0
    for row in sheet_headline.iter_rows():
        if row[0] is str:
            continue
        time1 = row[1].value
        tic1 = row[2].value
        if date == row[0].value and time1 > open_time and time1 < end_time:
            for row1 in equities:
                vol3m = None
                if row1[0] == tic1:
                    print(tic1)
                    print(row1)
                    vol3m = row1[3]
                    break
            count += 1
            data.append([count, row[2].value, row[10].value, None, None, time1, vol3m])
    data_copy = data
    if len(data) > 1:
        date_str = date.strftime("%Y-%m-%d")
        date_str = date_str.replace("-", "_")
        url_symbol = "http://finance.yahoo.com/q?s="
        for i, stock in enumerate(data):
            if i == 0:
                continue
            tic = stock[1]
            print("now processing ticker: ", tic)
            signal = stock[2]
            req = requests.get(url_symbol+tic)
            soup = BeautifulSoup(req.text, "html5lib")

            exchange1 = soup.find("span", class_='C($finDarkGray) Fz(12px)')
            try:
                exchange = exchange1.text
            except:
                print(tic, " cannot find exchange on yahoo")
                continue
            if exchange[0:2] == "Na":
                data_copy[i][1] = "NASDAQ:"+tic
            elif "NYSE MKT" in exchange:
                data_copy[i][1] = "NYSE MKT:"+tic
            elif exchange[0:2] == "NY":
                data_copy[i][1] = "NYSE:"+tic
            else:
                print("cant find exchange match on yehoo for ticker: ", tic)
            if signal == "SELL":
                data_copy[i][3] = -1
            elif signal == "BUY":
                data_copy[i][3] = 1
            else:
                sys.exit(" NO SELL/MATCH VALUE MATCHED for ticker at date", tic, date)

            datalist = GetPriceURL(tic)

            open_price = IntradayPrice(tic, date_str, str(stock[5]), datalist)
            next_min = 0
            if open_price == "NULL":
                next_min += 1
                cap1 = "NULL"
            elif stock[6] is None:
                cap1 = cap_ipo
            elif stock[6] == "N/A" or stock[6] == "":
                cap1 = "N/A"
            else:
                print(stock)
                cap1 = open_price * float(stock[6]) / division
                cap1 /= NAV
                print(cap1)
                if cap1 > cap:
                    cap1 = cap
            print("ticker,openprice,AvgVol3m,division,cap are: ", tic, open_price, stock[6], division, cap1)
            data_copy[i][4] = cap1

        path_save = path_output + "/" + date_str.replace("_", "") + "_" + name_output + ".csv"
        print("file saved to :" + path_save)
        data_copy = [x[0:6] for x in data_copy]
        with open(path_save, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(data_copy)
    else:
        print("WEEKEND!! no data in headline file at date: ", date)
