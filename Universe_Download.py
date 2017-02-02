import pandas as pd
import openpyxl


df1 = pd.read_csv('http://www.nasdaq.com/screening/companies-by-name.aspx?letter=0&exchange=nasdaq&render=download')
df2 = pd.read_csv('http://www.nasdaq.com/screening/companies-by-name.aspx?letter=0&exchange=nyse&render=download')
df3 = pd.read_csv('http://www.nasdaq.com/screening/companies-by-name.aspx?letter=0&exchange=amex&render=download')
df = pd.concat([df1, df2, df3])
df.index = range(len(df))
# df.to_csv('/Users/pengjiawei/Desktop/m_global/project4/Input/Company List.csv')
df['ADR'] = None
df = df.drop(['Sector', 'industry', 'Summary Quote', 'Unnamed: 8'], 1)
df = df[['Symbol', 'Name', 'LastSale', 'MarketCap', 'ADR', 'IPOyear']]
df['Exchange'] = 1
df['AvgVol10d'] = None
df['AvgVol3m'] = None
df['10d_Median_vol'] = None
df['60d_Median_vol'] = None
df['Beta3Y'] = None
df['SectorCode'] = None
df.loc[:len(df1), 'Exchange'] = 'NASDAQ'
df.loc[len(df1):len(df1)+len(df2), 'Exchange'] = 'NYSE'
df.loc[len(df1)+len(df2):len(df), 'Exchange'] = 'AMEX'
# df.to_csv('/Users/pengjiawei/Desktop/m_global/project4/Input/Combine.csv', index=False)

Config_Path = '/Users/pengjiawei/Desktop/m_global/config/Price & Volume config_NEW.xlsx'
Config_Sheetname = 'Tradable Universe Update'
wb = openpyxl.load_workbook(Config_Path)
ws = wb.get_sheet_by_name(Config_Sheetname)
file_path = ws['B2'].value
sheet_name = ws['B3'].value
out_path = ws['B4'].value
remove_path = ws['B5'].value
remove_sheet = ws['B6'].value
wb.save(Config_Path)
'''
remove_list = pd.read_excel(remove_path).Ticker.unique()
C_list = df['Symbol'].values.tolist()
for i in remove_list:
    if i in C_list:
        df = df.drop(C_list.index(i))
df.index = range(len(df))
df.to_excel(out_path)
'''